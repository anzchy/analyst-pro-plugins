"""Tests for merge_financials.py — run in conda `web-scrape`:

    source "$(conda info --base)/etc/profile.d/conda.sh" && conda activate web-scrape
    cd analyst-deal/scripts && python3 -m pytest test_merge_financials.py -q

Scope: this file covers every case that lives IN the merge script (plan test
plan #1-10, 13, 14, 15, 16, 17). Out of scope here, by design, because they
live in other work items:
  * #11 mtime cache guard, #12 filename period parser, #18 --extract-only,
    #19 folder-scoped cache, #20 period-collision  → WT-B (command layer)
  * #21 agent anti-fabrication wording             → WT-C (agent); the script
    side of it (null → skipped cell) IS tested here as #14.

Smoke-test finding F1 (agent drifts _meta.period_date → report_date) hardened
the script side of #21: _period_note now accepts the report_date synonym and
emits a visible NOTE when neither key is present. Covered by test_21* below.
"""

import csv
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest

import merge_financials as mf

SHEET = mf.SHEET


# --------------------------------------------------------------------------- #
# fixtures / helpers
# --------------------------------------------------------------------------- #
def make_xlsx(path: Path, header_dates, rows):
    """header_dates: list placed in row1 from col1 (col1 usually the '项目' label).
    rows: list of (label, [existing values aligned to header_dates[1:]])."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    for ci, h in enumerate(header_dates, start=1):
        ws.cell(row=1, column=ci).value = h
    for ri, (label, vals) in enumerate(rows, start=2):
        ws.cell(row=ri, column=1).value = label
        for ci, v in enumerate(vals, start=2):
            ws.cell(row=ri, column=ci).value = v
    wb.save(path)


def make_sidecar(path: Path, items, period_date="2025-12-31"):
    obj = {
        "_meta": {"schema": "fin-sidecar/v1", "company": "T",
                  "quarter": "2025Q4", "period_date": period_date, "unit": "万元"},
        "items": items,
    }
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")


def col(ws, c, r):
    return ws.cell(row=r, column=c).value


def load(path):
    return openpyxl.load_workbook(path)[SHEET]


# --------------------------------------------------------------------------- #
# 1. xlsx: TARGET > latest → append new column
# --------------------------------------------------------------------------- #
def test_01_xlsx_target_after_latest_appends(tmp_path):
    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2023, 12, 31), datetime(2024, 12, 31)],
              [("货币资金", [10, 20])])
    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 30.0})
    out = mf.run(str(x), str(j), "2025-12-31")
    ws = load(x)
    assert col(ws, 4, 1) == datetime(2025, 12, 31)
    assert col(ws, 4, 2) == 30.0
    assert col(ws, 2, 2) == 10 and col(ws, 3, 2) == 20  # old cols intact
    assert any("INSERT" in line for line in out)


# --------------------------------------------------------------------------- #
# 2. xlsx: TARGET == latest → overwrite, idempotent rerun
# --------------------------------------------------------------------------- #
def test_02_xlsx_overwrite_idempotent(tmp_path):
    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2024, 12, 31), datetime(2025, 12, 31)],
              [("货币资金", [10, 999])])
    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 30.0})

    mf.run(str(x), str(j), "2025-12-31")
    ws = load(x)
    assert ws.max_column == 3  # no new column
    assert col(ws, 3, 2) == 30.0

    mf.run(str(x), str(j), "2025-12-31")  # rerun
    ws2 = load(x)
    assert ws2.max_column == 3
    assert col(ws2, 3, 2) == 30.0  # identical → idempotent


# --------------------------------------------------------------------------- #
# 3 / 16. OV1: TARGET < latest → insert-in-order (backfill), NOT abort
# --------------------------------------------------------------------------- #
def test_03_ov1_backfill_inserts_in_order(tmp_path):
    x = tmp_path / "h.xlsx"
    # sheet already has 2025-12-31; an earlier 2024-12-31 arrives
    make_xlsx(x, ["项目", datetime(2025, 12, 31)], [("货币资金", [999])])
    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 30.0}, period_date="2024-12-31")
    out = mf.run(str(x), str(j), "2024-12-31")
    ws = load(x)
    # new earlier column inserted BEFORE the later one (chronological order)
    assert col(ws, 2, 1) == datetime(2024, 12, 31)
    assert col(ws, 2, 2) == 30.0
    assert col(ws, 3, 1) == datetime(2025, 12, 31)  # pushed right
    assert col(ws, 3, 2) == 999
    assert any("回插" in line for line in out)


def test_03b_ov1_backfill_between_two(tmp_path):
    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2023, 12, 31), datetime(2025, 12, 31)],
              [("货币资金", [11, 33])])
    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 22.0}, period_date="2024-12-31")
    mf.run(str(x), str(j), "2024-12-31")
    ws = load(x)
    assert [col(ws, c, 1) for c in (2, 3, 4)] == [
        datetime(2023, 12, 31), datetime(2024, 12, 31), datetime(2025, 12, 31)]
    assert [col(ws, c, 2) for c in (2, 3, 4)] == [11, 22.0, 33]


# --------------------------------------------------------------------------- #
# 4. classify()
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("label,expected", [
    ("科目", "ignore"), ("项目", "ignore"),
    ("一、经营活动产生的现金流量：", "ignore"),
    ("流动资产合计", "skip"), ("营业总成本小计", "skip"), ("毛利率", "skip"),
    ("净利润", "skip"), ("经营活动产生的现金流量净额", "skip"), ("毛利", "skip"),
    ("货币资金", "write"), ("应收账款净额", "write"), ("一、营业总收入", "write"),
])
def test_04_classify(label, expected):
    assert mf.classify(label) == expected


# --------------------------------------------------------------------------- #
# 5. label drift: sheet label ∉ json → reported missing, not silent
# --------------------------------------------------------------------------- #
def test_05_label_drift_reported(tmp_path):
    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2024, 12, 31)],
              [("货币资金", [1]), ("存货", [2])])
    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 30.0})  # 存货 absent
    out = mf.run(str(x), str(j), "2025-12-31")
    ws = load(x)
    assert col(ws, 3, 3) is None  # 存货 not written
    joined = "\n".join(out)
    assert "missing" in joined and "存货" in joined


# --------------------------------------------------------------------------- #
# 6. csv ↔ xlsx parity: same JSON → identical detail-row values
# --------------------------------------------------------------------------- #
def test_06_csv_xlsx_parity(tmp_path):
    labels = [("货币资金", "write"), ("流动资产合计", "skip"),
              ("应收账款净额", "write")]
    items = {"货币资金": 12.5, "流动资产合计": 88.0, "应收账款净额": 7.25}

    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2024, 12, 31)], [(l, [0]) for l, _ in labels])
    c = tmp_path / "h.csv"
    with open(c, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["项目", "2024-12-31"])
        for l, _ in labels:
            w.writerow([l, "0"])

    j = tmp_path / "s.json"
    make_sidecar(j, items)
    mf.run(str(x), str(j), "2025-12-31")
    mf.run(str(c), str(j), "2025-12-31")

    ws = load(x)
    xlsx_vals = {col(ws, 1, r): col(ws, 3, r) for r in range(2, ws.max_row + 1)}
    with open(c, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    csv_vals = {r[0]: r[2] for r in rows[1:]}

    assert xlsx_vals["货币资金"] == 12.5
    assert float(csv_vals["货币资金"]) == 12.5
    assert xlsx_vals["应收账款净额"] == 7.25
    assert float(csv_vals["应收账款净额"]) == 7.25
    assert xlsx_vals["流动资产合计"] is None      # skip in both
    assert csv_vals["流动资产合计"] == ""


# --------------------------------------------------------------------------- #
# 7. G1: malformed / empty / missing / schema-less side-file → clean error
# --------------------------------------------------------------------------- #
def test_07_sidecar_errors_are_clean(tmp_path):
    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2024, 12, 31)], [("货币资金", [1])])

    bad = tmp_path / "bad.json"
    bad.write_text("{ not json", encoding="utf-8")
    empty = tmp_path / "empty.json"
    empty.write_text("", encoding="utf-8")
    noitems = tmp_path / "noitems.json"
    noitems.write_text('{"_meta": {}}', encoding="utf-8")

    for p in (bad, empty, noitems, tmp_path / "nope.json"):
        with pytest.raises(mf.MergeError):
            mf.run(str(x), str(p), "2025-12-31")

    # main() converts MergeError → exit 1, message on stderr, no traceback
    r = subprocess.run(
        [sys.executable, str(Path(__file__).parent / "merge_financials.py"),
         "--target", str(x), "--json", str(bad), "--date", "2025-12-31"],
        capture_output=True, text=True)
    assert r.returncode == 1
    assert "Traceback" not in r.stderr
    assert r.stderr.startswith("ERR:")


# --------------------------------------------------------------------------- #
# 8. G2: csv encoding matrix (plain utf-8, utf-8-BOM, Chinese labels)
# --------------------------------------------------------------------------- #
@pytest.mark.parametrize("enc", ["utf-8", "utf-8-sig"])
def test_08_csv_encoding_roundtrip(tmp_path, enc):
    c = tmp_path / "h.csv"
    with open(c, "w", encoding=enc, newline="") as f:
        w = csv.writer(f)
        w.writerow(["项目", "2024-12-31"])
        w.writerow(["货币资金", "1"])
        w.writerow(["应收账款净额", "2"])
    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 9.9, "应收账款净额": 8.8})
    mf.run(str(c), str(j), "2025-12-31")
    with open(c, encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    assert rows[0] == ["项目", "2024-12-31", "2025-12-31"]
    body = {r[0]: r[2] for r in rows[1:]}
    assert float(body["货币资金"]) == 9.9
    assert float(body["应收账款净额"]) == 8.8


# --------------------------------------------------------------------------- #
# 9. G2: csv with no YYYY-MM-DD header → clean error
# --------------------------------------------------------------------------- #
def test_09_csv_no_date_header_errors(tmp_path):
    c = tmp_path / "h.csv"
    with open(c, "w", encoding="utf-8", newline="") as f:
        csv.writer(f).writerows([["项目", "2024年报"], ["货币资金", "1"]])
    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 1.0})
    with pytest.raises(mf.MergeError, match="日期表头"):
        mf.run(str(c), str(j), "2025-12-31")


# --------------------------------------------------------------------------- #
# 10. G3: extension dispatch
# --------------------------------------------------------------------------- #
def test_10_extension_dispatch(tmp_path):
    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 1.0})

    up = tmp_path / "H.XLSX"
    make_xlsx(up, ["项目", datetime(2024, 12, 31)], [("货币资金", [1])])
    mf.run(str(up), str(j), "2025-12-31")  # uppercase .XLSX → xlsx branch ok

    for name in ("h.xls", "h.txt", "h"):
        p = tmp_path / name
        p.write_text("x", encoding="utf-8")
        with pytest.raises(mf.MergeError):
            mf.run(str(p), str(j), "2025-12-31")

    with pytest.raises(mf.MergeError, match="--date"):
        mf.run(str(up), str(j), "2025-13-99")  # bad date is clean too


# --------------------------------------------------------------------------- #
# 13. edge: two row-1 cells == latest → deterministic (leftmost) target column
# --------------------------------------------------------------------------- #
def test_13_duplicate_latest_deterministic(tmp_path):
    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2024, 12, 31), datetime(2024, 12, 31)],
              [("货币资金", [111, 222])])
    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 30.0})
    mf.run(str(x), str(j), "2024-12-31")
    ws = load(x)
    assert col(ws, 2, 2) == 30.0   # leftmost duplicate overwritten
    assert col(ws, 3, 2) == 222    # other left untouched


# --------------------------------------------------------------------------- #
# 14. edge: JSON value null → cell skipped (never "None"/""/0)
# --------------------------------------------------------------------------- #
def test_14_null_value_skips_cell(tmp_path):
    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2024, 12, 31)],
              [("货币资金", [1]), ("应收账款净额", [2])])
    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 5.0, "应收账款净额": None})
    mf.run(str(x), str(j), "2025-12-31")
    ws = load(x)
    assert col(ws, 3, 2) == 5.0
    assert col(ws, 3, 3) is None  # null → genuinely empty, not "None"


# --------------------------------------------------------------------------- #
# 15. CRITICAL REGRESSION — golden fixture mirroring the real 三大财务报表 sheet
# --------------------------------------------------------------------------- #
def test_15_golden_regression(tmp_path):
    x = tmp_path / "历年财务报表.xlsx"
    make_xlsx(
        x,
        ["项目", datetime(2023, 12, 31), datetime(2024, 12, 31)],
        [
            ("货币资金", [100, 110]),                       # write
            ("应收账款净额", [50, 55]),                      # write
            ("流动资产合计", [200, 220]),                    # skip (合计)
            ("一、营业总收入", [300, 330]),                  # write
            ("毛利", [120, 130]),                            # skip (SKIP_EXACT)
            ("毛利率", [40.0, 39.4]),                        # skip (率)
            ("净利润", [80, 88]),                            # skip (SKIP_EXACT)
            ("一、经营活动产生的现金流量：", [None, None]),   # ignore (SILENT)
            ("经营活动产生的现金流量净额", [60, 66]),         # skip (SKIP_EXACT)
            ("销售商品、提供劳务收到的现金", [310, 340]),     # write
            ("存货", [70, 77]),                              # missing (drift)
        ],
    )
    j = tmp_path / "s.json"
    make_sidecar(j, {
        "货币资金": 121.0,
        "应收账款净额": 60.5,
        "流动资产合计": 242.0,           # present but must be SKIPPED by sheet rule
        "一、营业总收入": 363.0,
        "毛利率": 38.0,                  # present but skipped
        "净利润": 96.0,                  # present but skipped
        "经营活动产生的现金流量净额": 72.0,
        "销售商品、提供劳务收到的现金": 374.0,
    })
    out = mf.run(str(x), str(j), "2025-12-31")
    ws = load(x)

    assert col(ws, 4, 1) == datetime(2025, 12, 31)          # appended col D
    # detail rows written
    assert col(ws, 4, 2) == 121.0                            # 货币资金
    assert col(ws, 4, 3) == 60.5                             # 应收账款净额
    assert col(ws, 4, 5) == 363.0                            # 一、营业总收入
    assert col(ws, 4, 11) == 374.0                           # 销售...收到的现金
    # classification rows must stay empty (formula/structural — Step 5.5.2 rule)
    assert col(ws, 4, 4) is None                             # 流动资产合计 skip
    assert col(ws, 4, 6) is None                             # 毛利 skip
    assert col(ws, 4, 7) is None                             # 毛利率 skip
    assert col(ws, 4, 8) is None                             # 净利润 skip
    assert col(ws, 4, 9) is None                             # 现金流分段标签 ignore
    assert col(ws, 4, 10) is None                            # 净额 skip
    assert col(ws, 4, 12) is None                            # 存货 missing → empty
    # historical columns untouched
    assert col(ws, 2, 2) == 100 and col(ws, 3, 2) == 110
    # label drift surfaced, not silent
    joined = "\n".join(out)
    assert "存货" in joined and "missing" in joined
    # ignore label NOT counted as missing
    assert "一、经营活动产生的现金流量：" not in joined


# --------------------------------------------------------------------------- #
# 17. REGRESSION: quarterly-append shifts formula columns right, unchanged
# --------------------------------------------------------------------------- #
def test_17_quarterly_append_preserves_formula_column(tmp_path):
    x = tmp_path / "h.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET
    ws.cell(row=1, column=1).value = "项目"
    ws.cell(row=1, column=2).value = datetime(2025, 3, 31)
    ws.cell(row=1, column=3).value = datetime(2025, 6, 30)
    ws.cell(row=1, column=4).value = "占比"
    ws.cell(row=2, column=1).value = "货币资金"
    ws.cell(row=2, column=2).value = 10
    ws.cell(row=2, column=3).value = 20
    ws.cell(row=2, column=4).value = "=C2/SUM(C:C)"
    wb.save(x)

    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 30.0}, period_date="2025-09-30")
    mf.run(str(x), str(j), "2025-09-30")  # newer than all → append at col 4

    ws = load(x)
    assert col(ws, 4, 1) == datetime(2025, 9, 30)   # new period column
    assert col(ws, 4, 2) == 30.0
    assert col(ws, 5, 1) == "占比"                   # formula header shifted right
    assert col(ws, 5, 2) == "=C2/SUM(C:C)"           # formula text unchanged


# --------------------------------------------------------------------------- #
# 21. F1 (smoke-test finding): agent drifts _meta.period_date → report_date.
#     Cross-check must survive the synonym, and never vanish silently when
#     neither key is present.
# --------------------------------------------------------------------------- #
def test_21_period_note_drift_resilient():
    t = datetime(2025, 12, 31)
    # contract key, matches --date → silent OK
    assert mf._period_note({"period_date": "2025-12-31"}, t) == []
    # drifted to report_date, matches → accepted via synonym, silent OK
    assert mf._period_note({"report_date": "2025-12-31"}, t) == []
    # drifted to report_date, mismatched → mismatch NOTE still fires
    note = mf._period_note({"report_date": "2024-09-30"}, t)
    assert note and "不一致" in note[0]
    # neither key (deeper drift / extra-keys-only) → VISIBLE note, not silent
    note = mf._period_note({"source_file": "x.pdf"}, t)
    assert note and "缺 period_date" in note[0]


def test_21b_run_survives_report_date_drift(tmp_path):
    """End-to-end: a sidecar carrying the real F1 drift (report_date + extra
    keys, no period_date) still merges; the merge is driven by --date+items."""
    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2024, 12, 31)], [("货币资金", [1])])
    j = tmp_path / "s.json"
    j.write_text(json.dumps({
        "_meta": {"schema": "fin-sidecar/v1", "company": "矽昌通信",
                  "quarter": "2025Q4", "report_date": "2025-12-31",
                  "unit": "万元", "source_file": "x.pdf",
                  "pages_read": "1-3", "generated_by": "financial-analyzer"},
        "items": {"货币资金": 343.09},
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    out = mf.run(str(x), str(j), "2025-12-31")
    ws = load(x)
    assert col(ws, 3, 2) == 343.09          # real value still merged
    # report_date == --date → synonym accepted, no spurious mismatch NOTE
    assert not any("不一致" in line for line in out)


# --------------------------------------------------------------------------- #
# 22. STRUCTURAL FIX — full/half-width punctuation drift no longer drops a row.
#     Mirrors the real 矽昌通信 drift (sheet half-width vs side-file full-width)
#     that previously needed 7 hand-renamed sidecar keys.
# --------------------------------------------------------------------------- #
def test_22_punctuation_drift_now_matches(tmp_path):
    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2024, 12, 31)], [
        ("实收资本(或股本)", [1]),                       # half-width in sheet
        ('资产处置收益(损失以"-"号填列)', [2]),          # half-width parens+quotes
        ("加:营业外收入", [3]),                          # half-width colon
    ])
    j = tmp_path / "s.json"
    make_sidecar(j, {
        "实收资本（或股本）": 11.0,                       # full-width in side-file
        '资产处置收益（损失以"-"号填列）': 5.92,          # full-width parens+quotes
        "加：营业外收入": 0.0,                            # full-width colon
    })
    out = mf.run(str(x), str(j), "2025-12-31")
    ws = load(x)
    assert col(ws, 3, 2) == 11.0
    assert col(ws, 3, 3) == 5.92
    assert col(ws, 3, 4) == 0.0                           # 0.0 still written
    joined = "\n".join(out)
    assert "missing" in joined  # the line is present...
    assert "Missing labels:" not in joined                # ...but count is 0
    assert "归一化匹配 = 3" in joined                    # surfaced, not silent


def test_22b_classify_punctuation_resilient():
    # full-width and half-width colon variants both stay structural/ignore
    assert mf.classify("一、经营活动产生的现金流量：") == "ignore"
    assert mf.classify("一、经营活动产生的现金流量:") == "ignore"
    # canonicalization helpers
    assert mf.normalize_label("实收资本（或股本）") == "实收资本(或股本)"
    assert mf.canon_label("减：营业总成本") == mf.canon_label("二、营业总成本")


# --------------------------------------------------------------------------- #
# 23. semantic alias (NOT punctuation): 减：营业总成本 ⇄ 二、营业总成本
# --------------------------------------------------------------------------- #
def test_23_alias_matches(tmp_path):
    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2024, 12, 31)], [("二、营业总成本", [9])])
    j = tmp_path / "s.json"
    make_sidecar(j, {"减：营业总成本": 497.06})
    out = mf.run(str(x), str(j), "2025-12-31")
    ws = load(x)
    assert col(ws, 3, 2) == 497.06
    assert "归一化匹配 = 1" in "\n".join(out)


# --------------------------------------------------------------------------- #
# 24. zero-fabrication: two side-file keys → same canon, different values →
#     surfaced as AMBIGUOUS and NOT written (never guessed).
# --------------------------------------------------------------------------- #
def test_24_ambiguous_collision_not_fabricated(tmp_path):
    x = tmp_path / "h.xlsx"
    # sheet label equals NEITHER raw key (so exact path can't resolve it) but
    # canon-collides with both → must refuse to guess.
    make_xlsx(x, ["项目", datetime(2024, 12, 31)], [("投资收益（损失）", [1])])
    j = tmp_path / "s.json"
    make_sidecar(j, {"投资收益(损失)": 3.0, "投资收益（损失）　": 7.0})
    out = mf.run(str(x), str(j), "2025-12-31")
    ws = load(x)
    assert col(ws, 3, 2) is None                          # nothing written
    joined = "\n".join(out)
    assert "AMBIGUOUS" in joined
    assert "投资收益（损失）" in joined


# --------------------------------------------------------------------------- #
# 25. genuinely-absent label still reported missing (fix must NOT mask gaps)
# --------------------------------------------------------------------------- #
def test_25_genuine_absence_still_missing(tmp_path):
    x = tmp_path / "h.xlsx"
    make_xlsx(x, ["项目", datetime(2024, 12, 31)],
              [("货币资金", [1]), ("存货", [2])])
    j = tmp_path / "s.json"
    make_sidecar(j, {"货币资金": 30.0})                   # 存货 truly absent
    out = mf.run(str(x), str(j), "2025-12-31")
    ws = load(x)
    assert col(ws, 3, 3) is None
    joined = "\n".join(out)
    assert "Missing labels:" in joined and "存货" in joined
    assert "归一化匹配 = 0" in joined                    # not a false normalize
