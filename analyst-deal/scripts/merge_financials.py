#!/usr/bin/env python3
"""Merge one reporting period's extracted financials into a historical table.

Extracted from `analyst-deal/commands/portfolio-tracking.md` Step 5.5.2 so the
standalone `/analyst-deal:financial-analyzer` command and `portfolio-tracking`
share ONE merge implementation (no copy-paste). Contract is frozen in
`docs/designs/fin-sidecar-contract.md` — read it before changing anything here.

Must run inside the conda `web-scrape` env (openpyxl lives there; see
`analyst-deal/CLAUDE.md`).

CLI:
    python3 merge_financials.py --target <xlsx|csv> --json <sidecar.json> \
                                --date <YYYY-MM-DD>

Differences from the original inline Step 5.5.2 (locked by /plan-eng-review):
  * A2  side-file is stdlib JSON, not YAML (PyYAML absent in user env).
  * OV1 `TARGET < latest` no longer aborts — it inserts the column in
        chronological order (backfill is a normal multi-PDF workflow).
  * A4  a stdlib-csv branch shares the exact classify()+value pipeline with
        the xlsx branch (parity-tested).
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from datetime import datetime

SHEET = "三大财务报表"

# --- Row classification: copied VERBATIM from portfolio-tracking Step 5.5.2.
#     Do not edit these constants without re-running the golden regression test.
SILENT_IGNORE = {
    "科目", "项目",                              # 列头
    "利润表", "现金流量表",                       # sheet 内部分段标签
    "一、经营活动产生的现金流量：",
    "二、投资活动产生的现金流量：",
    "三、筹资活动产生的现金流量：",
}
SKIP_EXACT = {
    # 利润表 derived
    "毛利", "营业利润", "利润总额", "净利润",
    # 现金流量表 净额 (= 流入小计 - 流出小计)
    "经营活动产生的现金流量净额",
    "投资活动产生的现金流量净额",
    "筹资活动产生的现金流量净额",
    # 现金流量表合算行
    "五、现金及现金等价物净增加额",
    "六、期末现金及现金等价物余额",
}


# --------------------------------------------------------------------------- #
# Label canonicalization — for MATCHING ONLY. Display / reporting always keeps
# the ORIGINAL label.
#
# Why: financial-analyzer extracts 合并报表 行项 with official full-width CJK
# punctuation （）：，"" ; hand-maintained 历年表 first-column labels often use
# half-width ()：,"" . The old exact-string match silently demoted a populated
# detail row to "missing" on a single full/half-width slip (the exact drift the
# report's `missing` column warns about — observed in the 矽昌通信 run, where 7
# sidecar keys had to be hand-renamed). Folding punctuation + a tiny auditable
# synonym table fixes this structurally so the fix survives plugin upgrades and
# no longer needs per-run manual key surgery. Genuinely-absent labels are still
# reported missing; ambiguous synonym collisions are surfaced, never guessed.
# --------------------------------------------------------------------------- #
_PUNCT_FOLD = str.maketrans({
    "（": "(", "）": ")",
    "：": ":",
    "，": ",",
    "　": " ",              # U+3000 ideographic space
    "“": '"', "”": '"',   # “ ” → "
    "‘": "'", "’": "'",   # ‘ ’ → '
})


def normalize_label(label: str) -> str:
    """Punctuation-folded, outer-whitespace-stripped form used only to match."""
    return label.translate(_PUNCT_FOLD).strip()


# Known SEMANTIC synonyms (NOT mere punctuation): reading a 利润表 PDF the agent
# may emit the standard "减：营业总成本" while a hand-kept 历年表 row reads
# "二、营业总成本" — the same line. Add a group here (raw forms) ONLY for a
# real, recurring synonym observed in practice. This is the single place a
# non-identity label mapping is allowed — keep it short and auditable.
LABEL_ALIASES: list[set[str]] = [
    {"二、营业总成本", "减：营业总成本"},
]

_ALIAS_CANON: dict[str, str] = {}
for _grp in LABEL_ALIASES:
    _members = sorted(normalize_label(m) for m in _grp)
    for _m in _members:
        _ALIAS_CANON[_m] = _members[0]


def canon_label(label: str) -> str:
    """normalize_label + collapse a known synonym group to one representative."""
    n = normalize_label(label)
    return _ALIAS_CANON.get(n, n)


# Constant sets compared in canonical space so a full/half-width variant of a
# structural / subtotal label still classifies correctly (e.g. a sheet using
# `一、经营活动产生的现金流量:` with a half-width colon stays `ignore`).
_SILENT_IGNORE_N = {normalize_label(x) for x in SILENT_IGNORE}
_SKIP_EXACT_N = {normalize_label(x) for x in SKIP_EXACT}


def classify(label: str) -> str:
    """ignore (structural, not missing) / skip (subtotal·ratio) / write (detail).

    Compared on the canonical form so punctuation drift cannot reclassify a row.
    """
    n = normalize_label(label)
    if n in _SILENT_IGNORE_N:
        return "ignore"
    if n.endswith("合计") or n.endswith("小计") or n.endswith("率"):
        return "skip"
    if n in _SKIP_EXACT_N:
        return "skip"
    return "write"


class MergeError(Exception):
    """Clean, user-facing failure. main() prints the message and exits 1 —
    callers never see a Python traceback (contract §2)."""


# --------------------------------------------------------------------------- #
# Side-file (JSON v1)
# --------------------------------------------------------------------------- #
def load_sidecar(path: str) -> tuple[dict, dict]:
    """Return (items, meta). Every failure path raises MergeError, never a
    bare json/IO traceback (test G1)."""
    try:
        with open(path, encoding="utf-8") as f:
            obj = json.load(f)
    except FileNotFoundError:
        raise MergeError(f"ERR: side-file 不存在: {path}")
    except json.JSONDecodeError as e:
        raise MergeError(f"ERR: side-file 不是合法 JSON ({path}): {e}")
    except OSError as e:
        raise MergeError(f"ERR: 无法读取 side-file {path}: {e}")

    if not isinstance(obj, dict) or not isinstance(obj.get("items"), dict):
        raise MergeError(
            f"ERR: side-file 缺少 'items' 对象（格式不符 fin-sidecar/v1 或文件损坏）: {path}"
        )
    meta = obj.get("_meta") if isinstance(obj.get("_meta"), dict) else {}
    return obj["items"], meta


# --------------------------------------------------------------------------- #
# Shared placement decision (OV1) + shared write pipeline (A4)
# --------------------------------------------------------------------------- #
def decide_placement(anchors: list[tuple[int, datetime]], target: datetime) -> tuple[str, int]:
    """anchors: (column_index, date) for every period column, ascending column order.

    Returns (mode, column):
      ('overwrite', c)      target already present → clear+rewrite column c (idempotent)
      ('insert_before', c)  insert a new column at c, before the earliest column later
                            than target (chronological backfill — OV1)
      ('append', c)         target newer than all → new column right of the latest
                            (preserves pre-extraction quarterly-append behavior)
    Ties (two columns == target) resolve to the leftmost — deterministic (test 13).
    """
    exact = [c for c, d in anchors if d == target]
    if exact:
        return ("overwrite", min(exact))
    later = [(c, d) for c, d in anchors if d > target]
    if later:
        col = min(later, key=lambda cd: (cd[1], cd[0]))[0]
        return ("insert_before", col)
    latest_col = max(anchors, key=lambda cd: (cd[1], cd[0]))[0]
    return ("append", latest_col + 1)


def _index_items(items: dict) -> tuple[dict, dict]:
    """Build canon-key → (raw_key, value). Two DIFFERENT raw keys collapsing to
    the same canon form with DIFFERENT values are ambiguous — recorded so the
    canonical fallback never silently guesses between synonyms (zero-fab)."""
    idx: dict[str, tuple[str, object]] = {}
    ambiguous: dict[str, set[str]] = {}
    for k, v in items.items():
        ck = canon_label(k)
        if ck in idx:
            prev_k, prev_v = idx[ck]
            if k != prev_k and v != prev_v:
                ambiguous.setdefault(ck, {prev_k}).add(k)
            continue  # keep first seen; an exact per-label hit still wins below
        idx[ck] = (k, v)
    return idx, ambiguous


def plan_writes(labels: list[tuple[int, object]], items: dict):
    """The single source of truth for classify()+null handling — both the xlsx
    and csv branches call this so their detail-row values are identical by
    construction (parity test #6).

    Matching order per label: exact (contract's primary, fast path) → canonical
    (punctuation/alias-insensitive fallback). Canonical hits are recorded in
    `audit["normalized"]` (surfaced, not silent — same transparency rule as
    `missing`); unresolved synonym collisions go to `audit["ambiguous"]` and
    are NOT written.

    labels: (row_id, first_column_label) in sheet order.
    Returns (written, skipped, missing, writes, audit) where
    writes = [(row_id, value)] and
    audit = {"normalized": [(rid, sheet_label, sidecar_key)],
             "ambiguous":  [(rid, sheet_label, [candidate_keys])]}.
    """
    written, skipped, missing, writes = [], [], [], []
    normalized: list[tuple[int, str, str]] = []
    ambiguous_hit: list[tuple[int, str, list[str]]] = []
    canon_idx, ambiguous = _index_items(items)
    for rid, label in labels:
        if not label or not isinstance(label, str):
            continue
        cls = classify(label)
        if cls == "ignore":
            continue
        if cls == "skip":
            skipped.append((rid, label))
            continue
        # cls == "write"
        if label in items:                       # exact — primary contract match
            val, via = items[label], None
        else:
            ck = canon_label(label)
            if ck in ambiguous:                  # synonym collision → never guess
                ambiguous_hit.append((rid, label, sorted(ambiguous[ck])))
                continue
            hit = canon_idx.get(ck)              # punctuation/alias-insensitive
            if hit is None:
                missing.append((rid, label))
                continue
            val, via = hit[1], hit[0]
        if val is None:
            continue  # explicit null → skip cell, never "None"/"" /0 (test 14)
        writes.append((rid, val))
        written.append((rid, label))
        if via is not None:
            normalized.append((rid, label, via))
    audit = {"normalized": normalized, "ambiguous": ambiguous_hit}
    return written, skipped, missing, writes, audit


def _col_letter(c: int) -> str:
    s = ""
    while c > 0:
        c, rem = divmod(c - 1, 26)
        s = chr(65 + rem) + s
    return s


def _period_note(meta: dict, target: datetime) -> list[str]:
    # Contract requires _meta.period_date. The agent is an LLM and is known to
    # drift the key to report_date (smoke-test finding F1). Accept that one
    # synonym so the cross-check survives drift; if NEITHER key is present the
    # cross-check would otherwise vanish silently — emit a visible NOTE instead.
    pd = meta.get("period_date") or meta.get("report_date")
    if pd is None:
        return [
            "NOTE: side-file _meta 缺 period_date（也无 report_date）"
            "，跳过报告期交叉校验 — agent 可能键名 drift，请人工核对该期日期"
        ]
    if pd != target.strftime("%Y-%m-%d"):
        return [
            f"NOTE: side-file _meta.period_date={pd} 与 --date={target.date()} 不一致"
            f"（按 --date 写入；如非有意重跑请核对）"
        ]
    return []


# --------------------------------------------------------------------------- #
# xlsx branch — extracted Step 5.5.2
# --------------------------------------------------------------------------- #
def merge_xlsx(target: str, items: dict, meta: dict, target_dt: datetime) -> list[str]:
    import openpyxl  # imported lazily so csv-only use never needs it

    try:
        wb = openpyxl.load_workbook(target, data_only=False)
    except FileNotFoundError:
        raise MergeError(f"ERR: 历年表不存在: {target}")
    except Exception as e:  # openpyxl raises BadZipFile etc. on a non-xlsx
        raise MergeError(f"ERR: 无法打开 xlsx {target}: {e}")

    if SHEET not in wb.sheetnames:
        raise MergeError(
            f"FATAL: sheet {SHEET!r} 不存在；现有 sheets={wb.sheetnames}"
        )
    ws = wb[SHEET]

    anchors = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if isinstance(v, datetime):
            anchors.append((c, v))
    if not anchors:
        raise MergeError("FATAL: row 1 中没有任何 datetime cell；无法定位插入位置")

    mode, target_col = decide_placement(anchors, target_dt)
    out = _period_note(meta, target_dt)

    if mode == "overwrite":
        for r in range(2, ws.max_row + 1):
            ws.cell(row=r, column=target_col).value = None
        out.append(
            f"OVERWRITE: TARGET={target_dt.date()} 已存在于第 {target_col} 列，"
            f"清空旧值后覆盖（重跑场景）"
        )
    elif mode == "insert_before":
        ws.insert_cols(target_col)
        out.append(
            f"INSERT: 在第 {target_col} 列处按报告期顺序回插新列"
            f"（其后列右移）"
        )
    else:  # append
        ws.insert_cols(target_col)
        out.append(f"INSERT: 在第 {target_col} 列处追加新列（原列右移）")

    ws.cell(row=1, column=target_col).value = target_dt  # header as datetime

    labels = [
        (r, ws.cell(row=r, column=1).value) for r in range(2, ws.max_row + 1)
    ]
    written, skipped, missing, writes, audit = plan_writes(labels, items)
    for r, val in writes:
        ws.cell(row=r, column=target_col).value = val

    try:
        wb.save(target)
    except PermissionError:
        raise MergeError(
            f"FATAL: 无法写入 {target}（文件可能在 Excel 中打开）；请关闭后重跑。"
        )

    out += _report(target, target_col, target_dt, written, skipped, missing,
                    audit, letter=True)
    return out


# --------------------------------------------------------------------------- #
# csv branch — same classify()+value pipeline, no formulas
# --------------------------------------------------------------------------- #
def merge_csv(target: str, items: dict, meta: dict, target_dt: datetime) -> list[str]:
    try:
        with open(target, encoding="utf-8-sig", newline="") as f:
            rows = list(csv.reader(f))
    except FileNotFoundError:
        raise MergeError(f"ERR: 历年表不存在: {target}")
    except OSError as e:
        raise MergeError(f"ERR: 无法读取 csv {target}: {e}")

    if not rows or not rows[0]:
        raise MergeError(f"ERR: csv 为空或缺表头行: {target}")

    width = max(len(r) for r in rows)
    rows = [r + [""] * (width - len(r)) for r in rows]  # normalize ragged rows

    anchors = []
    for ci, cell in enumerate(rows[0]):
        try:
            d = datetime.strptime(cell.strip(), "%Y-%m-%d")
        except ValueError:
            continue  # label column / non-date header is simply not an anchor
        anchors.append((ci, d))
    if not anchors:
        raise MergeError(
            "ERR: csv row 0 没有任何 YYYY-MM-DD 日期表头；无法定位期间列"
        )

    mode, target_col = decide_placement(anchors, target_dt)
    out = _period_note(meta, target_dt)

    if mode == "overwrite":
        for r in rows[1:]:
            r[target_col] = ""
        out.append(
            f"OVERWRITE: TARGET={target_dt.date()} 已存在于第 {target_col + 1} 列，"
            f"清空旧值后覆盖（重跑场景）"
        )
    else:  # insert_before / append — same structural op for a csv
        for r in rows:
            r.insert(target_col, "")
        verb = "按报告期顺序回插" if mode == "insert_before" else "追加"
        out.append(f"INSERT: 在第 {target_col + 1} 列处{verb}新列（其后列右移）")

    rows[0][target_col] = target_dt.strftime("%Y-%m-%d")

    labels = [(ri, rows[ri][0]) for ri in range(1, len(rows))]
    written, skipped, missing, writes, audit = plan_writes(labels, items)
    for ri, val in writes:
        rows[ri][target_col] = "" if val is None else str(val)

    try:
        with open(target, "w", encoding="utf-8-sig", newline="") as f:
            csv.writer(f).writerows(rows)
    except OSError as e:
        raise MergeError(f"ERR: 无法写回 csv {target}: {e}")

    out += _report(target, target_col + 1, target_dt, written, skipped, missing,
                    audit, letter=False)
    out.append(
        "NOTE: csv 无公式 — 小计/比率行留空，需手工补或转 xlsx 用 SUM 公式"
    )
    return out


def _report(target, col, target_dt, written, skipped, missing, audit, *, letter):
    head = f"{_col_letter(col)} (idx={col})" if letter else f"第 {col} 列"
    norm = audit.get("normalized", []) if audit else []
    amb = audit.get("ambiguous", []) if audit else []
    lines = [
        f"OK: {target} 已更新",
        f"  新列 {head} = {target_dt.date()}",
        f"  written(detail/non-subtotal) = {len(written)}",
        f"    其中经全/半角或别名归一化匹配 = {len(norm)}（已自动对齐，非缺口）",
        f"  skipped(subtotal/ratio—需手工补或改 SUM 公式) = {len(skipped)}",
        f"  missing(label 在表但 side-file 中确实没有，非标点 drift)"
        f" = {len(missing)}",
    ]
    if missing:
        lines.append("  Missing labels:")
        for rid, lbl in missing:
            lines.append(f"    row {rid}: {lbl!r}")
    if norm:
        lines.append("  NOTE: 以下行经归一化后匹配 side-file 并写入（审计用）:")
        for rid, lbl, src in norm:
            lines.append(f"    row {rid}: 表标签 {lbl!r} ← side-file 键 {src!r}")
    if amb:
        lines.append(
            "  AMBIGUOUS: 多个 side-file 键归一化后同形且值不同 — 未写入，"
            "请人工消歧（zero-fab，绝不猜）:"
        )
        for rid, lbl, cands in amb:
            lines.append(f"    row {rid}: {lbl!r} ↔ 候选键 {cands!r}")
    return lines


# --------------------------------------------------------------------------- #
# CLI
# --------------------------------------------------------------------------- #
def run(target: str, json_path: str, date_str: str) -> list[str]:
    try:
        target_dt = datetime.strptime(date_str, "%Y-%m-%d")
    except ValueError:
        raise MergeError(f"ERR: --date 不是合法 YYYY-MM-DD: {date_str!r}")

    items, meta = load_sidecar(json_path)

    low = target.lower()
    if low.endswith(".xlsx"):
        return merge_xlsx(target, items, meta, target_dt)
    if low.endswith(".csv"):
        return merge_csv(target, items, meta, target_dt)
    if low.endswith(".xls"):
        raise MergeError(
            "ERR: 旧版 .xls 不支持；请在 Excel 中另存为 .xlsx 或导出 .csv 后重试"
        )
    raise MergeError(
        f"ERR: --target 扩展名无法识别（仅支持 .xlsx / .csv）: {target}"
    )


def main(argv=None) -> int:
    p = argparse.ArgumentParser(
        prog="merge_financials.py",
        description="把一期 JSON 侧文件并入历年财务报表 xlsx/csv（contract: docs/designs/fin-sidecar-contract.md）",
    )
    p.add_argument("--target", required=True, help="历年表 .xlsx 或 .csv")
    p.add_argument("--json", required=True, dest="json_path",
                   help="某期 fin-sidecar/v1 JSON 侧文件")
    p.add_argument("--date", required=True, help="报告期 YYYY-MM-DD")
    args = p.parse_args(argv)

    try:
        for line in run(args.target, args.json_path, args.date):
            print(line)
    except MergeError as e:
        print(str(e), file=sys.stderr)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
